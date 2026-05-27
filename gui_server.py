#!/usr/bin/env python3
"""
GUI Server — Interface Web para o Downloader slskd
===================================================
Servidor Flask que expõe uma API para a interface web controlar
o baixar_musicas.py em tempo real.

Instalar:
    pip install flask flask-cors

Uso:
    python3 gui_server.py
    Acesse: http://localhost:8082
"""

import json
import os
import re
import shutil
import sys
import threading
import time
import argparse
import logging
from datetime import datetime
from pathlib import Path
from queue import Queue, Empty
from io import BytesIO

try:
    from flask import Flask, jsonify, request, send_from_directory, Response
    from flask_cors import CORS
except ImportError:
    print("❌ Flask não encontrado. Instale com: pip install flask flask-cors")
    sys.exit(1)

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ─── Config padrão ────────────────────────────────────────────
SLSKD_API_KEY = "slskd123456"
SLSKD_URL      = "http://localhost:5030"
SLSKD_USUARIO  = "slskd"
SLSKD_SENHA    = "slskd"
LOG_DIR        = Path("logs")
DOWNLOADS_DIR  = Path.home() / "slskd" / "downloads"  # ~/slskd/downloads/


RAZAO_MINIMA = {"mp3": 1.8, "flac": 5.0, "wav": 5.0}
FORMATO_PRIORIDADE = {"flac": 3, "wav": 2, "mp3": 1}
# ──────────────────────────────────────────────────────────────

app    = Flask(__name__)
CORS(app)

# Estado global da aplicação
state = {
    "running":     False,
    "paused":      False,
    "stop_flag":   False,
    "total":       0,
    "baixados":    0,
    "erros":       0,
    "pendentes":   0,
    "atual":       "",
    "musicas":     [],
    "rejected_candidates": set(),
    "config": {
        "formatos":              ["flac", "wav", "mp3"],
        "bitrate_minimo":         192,
        "tempo_busca":           90,
        "tamanho_minimo":        7_000_000,
        "organizacao_modo":      "flat",  # flat, pasta_unica, artista, pasta_personalizada
        "pasta_download":        "",
        "pasta_personalizada":   "",
    },
}

log_queue = Queue()

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("gui")


# ─── Helpers ──────────────────────────────────────────────────

def push_log(msg: str, nivel: str = "info"):
    log_queue.put({"ts": datetime.now().strftime("%H:%M:%S"), "msg": msg, "nivel": nivel})


def criar_sessao():
    session = requests.Session()
    retry   = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def obter_token(session) -> bool:
    try:
        r = session.post(
            f"{SLSKD_URL}/api/v0/session",
            json={"username": SLSKD_USUARIO, "password": SLSKD_SENHA},
            timeout=10,
        )
        if r.status_code == 200:
            token = r.json().get("token")
            if token:
                session.headers.update({"Authorization": f"Bearer {token}"})
                return True
        push_log(f"Falha autenticação slskd: {r.status_code}", "error")
    except Exception as e:
        push_log(f"Erro conexão slskd: {e}", "error")
    return False


def limpar_termo(texto: str) -> str:
    texto = re.sub(r"$$.*?$$", "", texto)
    texto = re.sub(r"$$.*?$$", "", texto)
    texto = re.sub(r"\b(feat|ft|featuring)\.?\b.*", "", texto, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", texto).strip()


def calcular_razao_mb_min(tamanho: int, duracao: int) -> float:
    if duracao <= 0:
        return 0.0
    return (tamanho / 1_000_000) / (duracao / 60)


def calcular_score(c) -> float:
    score = c["prioridade"] * 10000
    if c["formato"] == "flac":  score += 5000
    elif c["formato"] == "wav": score += 3000
    elif c["formato"] == "mp3":
        br = c["bitrate"]
        if br >= 320:   score += 3200
        elif br >= 256: score += 2560
        elif br > 0:    score += br * 5
        else:           score += 1000
    score += c["tamanho"] / 1_000_000
    return score


def aguardar_resultados(session, search_id, tempo_max) -> list:
    """
    Aguarda a busca do slskd completar monitorando GET /api/v0/searches/{id}.
    Os arquivos só ficam disponíveis em /responses após state=Completed/TimedOut.
    """
    inicio = time.time()

    # Fase 1: aguarda sair do Queued
    push_log("⏳ Aguardando busca sair da fila...", "debug")
    while True:
        if time.time() - inicio > 60:
            push_log("⌛ Timeout aguardando fila — busca presa em Queued", "warn")
            return []
        try:
            r = session.get(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
            if r.status_code == 200:
                state_busca = r.json().get("state", "")
                if state_busca != "Queued":
                    push_log(f"▶ Busca iniciada (state={state_busca})", "debug")
                    break
        except Exception:
            pass
        time.sleep(2)

    # Fase 2: aguarda completar (fileCount sobe durante InProgress)
    ultimo_count = 0
    while True:
        elapsed = time.time() - inicio
        try:
            r = session.get(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
            if r.status_code == 200:
                dados      = r.json()
                is_done    = dados.get("isComplete", False)
                state_b    = dados.get("state", "")
                file_count = dados.get("fileCount", 0)

                if file_count != ultimo_count:
                    push_log(f"📦 Acumulando: {file_count} arquivos [{elapsed:.0f}s]", "debug")
                    ultimo_count = file_count

                # Busca concluída — agora busca os arquivos reais
                if is_done or "Completed" in state_b or "TimedOut" in state_b:
                    push_log(f"✅ Busca concluída: {file_count} arquivos ({elapsed:.0f}s)", "debug")
                    r2 = session.get(
                        f"{SLSKD_URL}/api/v0/searches/{search_id}/responses",
                        timeout=15,
                    )
                    if r2.status_code == 200:
                        respostas = r2.json() or []
                        total = sum(len(p.get("files", [])) for p in respostas)
                        push_log(f"📂 {total} arquivo(s) carregados de {len(respostas)} peer(s)", "debug")
                        return respostas
                    return []

                if elapsed > tempo_max:
                    push_log(f"⌛ Timeout {tempo_max}s (fileCount={file_count}) — aguardando Completed...", "warn")
                    # /responses só fica disponível após Completed; espera até 30s extras
                    deadline = time.time() + 30
                    while time.time() < deadline:
                        time.sleep(2)
                        try:
                            r_st = session.get(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
                            if r_st.status_code == 200:
                                d2 = r_st.json()
                                if d2.get("isComplete") or "Completed" in d2.get("state","") or "TimedOut" in d2.get("state",""):
                                    push_log(f"✅ Completed após timeout ({d2.get('fileCount')} arqs)", "debug")
                                    r2 = session.get(f"{SLSKD_URL}/api/v0/searches/{search_id}/responses", timeout=15)
                                    return r2.json() if r2.status_code == 200 else []
                        except Exception:
                            pass
                    push_log("⌛ Busca nunca completou — sem resultados", "warn")
                    return []

            elif r.status_code == 404:
                push_log("Search ID não encontrado (404)", "warn")
                return []
        except Exception:
            pass
        time.sleep(2)


def fazer_busca(session, termo, tempo_max) -> list:
    try:
        r = session.post(f"{SLSKD_URL}/api/v0/searches", json={"searchText": termo}, timeout=10)
        if r.status_code == 409:
            r_list = session.get(f"{SLSKD_URL}/api/v0/searches", timeout=10)
            if r_list.status_code == 200:
                for b in r_list.json():
                    if b.get("searchText", "").lower() == termo.lower():
                        return aguardar_resultados(session, b["id"], tempo_max)
            return []
        if r.status_code not in (200, 201):
            return []
        search_id = r.json().get("id")
        if not search_id:
            return []
        resultados = aguardar_resultados(session, search_id, tempo_max)
        # Deleta busca para não entupir a fila
        try:
            session.delete(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
        except Exception:
            pass
        return resultados
    except Exception as e:
        push_log(f"Erro busca: {e}", "error")
        return []


def buscar_musica(session, musica, artista, tempo_max) -> list:
    musica_limpa  = limpar_termo(musica)
    artista_limpo = limpar_termo(artista)
    primeiro      = artista_limpo.split(",")[0].split("&")[0].strip()
    tentativas_raw = [
        f"{artista_limpo} {musica_limpa}",
        f"{primeiro} {musica_limpa}",
        f"{primeiro} {' '.join(musica_limpa.split()[:3])}",
    ]
    vistas, tentativas = set(), []
    for t in tentativas_raw:
        if t.lower() not in vistas:
            vistas.add(t.lower())
            tentativas.append(t)

    for i, termo in enumerate(tentativas, 1):
        push_log(f"🔍 Tentativa {i}: {termo}")
        resultados = fazer_busca(session, termo, tempo_max)
        total = sum(len(r.get("files", [])) for r in resultados)
        if total > 0:
            push_log(f"📦 {total} arquivo(s) encontrado(s)")
            push_log(f"   Usuários: {[r.get('username', '') for r in resultados]}", "debug")
            return resultados
        else:
            push_log(f"   Nenhum resultado para este termo", "debug")

    push_log(f"⚠️ NENHUM resultado em nenhuma tentativa!", "warn")
    return []


def filtrar_candidatos(resultados, formatos, bitrate_min, tamanho_min, prioridade) -> list:
    candidatos = []
    rejeitados = []  # Para logs de debug

    # Log de debug: mostra usuários encontrados
    if resultados:
        usuarios_encontrados = [r.get('username', '') for r in resultados]
        push_log(f"   [DEBUG] Usuários na busca: {usuarios_encontrados}", "debug")

    for resp in resultados:
        usuario = resp.get("username", "")
        arquivos = resp.get("files", [])

        # Log: mostra arquivos de cada usuário antes de filtrar
        push_log(f"   [DEBUG] Usuario '{usuario}' tem {len(arquivos)} arquivo(s)", "debug")

        for arq in arquivos:
            nome    = arq.get("filename", "")
            ext     = nome.rsplit(".", 1)[-1].lower() if "." in nome else ""
            try:    bitrate = int(float(arq.get("bitRate") or arq.get("bitrate") or 0))
            except: bitrate = 0
            tamanho = arq.get("size", 0) or 0
            duracao = arq.get("length") or arq.get("duration") or 0
            razao   = calcular_razao_mb_min(tamanho, duracao)
            razao_min = RAZAO_MINIMA.get(ext, 0)

            # Log: mostra arquivo encontrado
            push_log(f"     [DEBUG] ARQUIVO: {nome[:50]} | ext={ext} size={tamanho/1e6:.1f}MB dur={duracao}s", "debug")

            # Verificações com logs detalhados
            if ext not in formatos:
                rejeitados.append({"nome": nome, "motivo": f"FORMATO={ext} (aceitos: {formatos})"})
                continue

            if "rejected_candidates" in state and (usuario, nome) in state["rejected_candidates"]:
                rejeitados.append({"nome": nome, "motivo": f"REJEITADO_ANTERIORMENTE por {usuario}"})
                continue

            if tamanho < tamanho_min:
                rejeitados.append({"nome": nome, "motivo": f"TAMANHO={tamanho/1e6:.1f}MB < {tamanho_min/1e6:.1f}MB"})
                continue

            if ext == "mp3" and bitrate > 0 and bitrate < bitrate_min:
                rejeitados.append({"nome": nome, "motivo": f"BITRATE={bitrate}kbps < {bitrate_min}kbps"})
                continue

            # RAZÃO: REMOVIDA TEMPORARIAMENTE PARA DEBUG
            # Era: if duracao > 0 and razao_min > 0 and razao < razao_min:
            # Agora: aceita tudo que passou nos filtros acima

            # ✅ Arquivo aceito
            c = {
                "usuario":    usuario,
                "arquivo":    nome,
                "formato":    ext,
                "bitrate":    bitrate,
                "tamanho":    tamanho,
                "duracao":    duracao,
                "razao":      razao,
                "prioridade": prioridade.get(ext, 0),
            }
            c["score"] = calcular_score(c)
            candidatos.append(c)

            push_log(f"  ✅ CANDIDATO: {nome} ({ext}, {tamanho/1e6:.1f}MB, {bitrate}kbps)", "debug")

    candidatos.sort(key=lambda x: x["score"], reverse=True)

    # Log dos arquivos rejeitados
    if rejeitados:
        push_log(f"  ⚠️  Rejeitados: {len(rejeitados)} arquivo(s)", "debug")
        for r in rejeitados[:5]:
            push_log(f"     - {r['nome'][:60]} → {r['motivo']}", "debug")
        if len(rejeitados) > 5:
            push_log(f"     ... e mais {len(rejeitados) - 5} arquivo(s)", "debug")

    push_log(f"  🎯 {len(candidatos)} candidato(s) válido(s) de {len(resultados)} resposta(s)", "debug")

    return candidatos


def sanitizar(nome: str) -> str:
    """Remove caracteres especiais para nome de pasta/arquivo"""
    if not nome:
        return "Desconhecido"
    for c in r'\/:*?"<>|':
        nome = nome.replace(c, "_")
    nome = nome.strip()
    if len(nome) > 100:
        nome = nome[:100]
    return nome or "Desconhecido"


def mover_arquivo_para_destino(nome_arquivo: str, artista: str, musica: str, organizacao_modo: str, pasta_personalizada: str = "") -> dict:
    """
    Move o arquivo baixado para o destino correto baseado no modo de organização.

    Retorna dict com:
        - success: bool
        - source: str (caminho original)
        - destination: str (novo caminho)
        - message: str (mensagem de status)
    """
    resultado = {"success": False, "source": "", "destination": "", "message": ""}

    try:
        # Busca recursiva do arquivo na pasta de downloads
        arquivo_encontrado = None
        for ext in ["", ".mp3", ".flac", ".wav"]:
            busca_nome = nome_arquivo if ext == "" else nome_arquivo
            for f in DOWNLOADS_DIR.rglob(f"*{ext}"):
                if f.name == nome_arquivo or nome_arquivo in f.name:
                    arquivo_encontrado = f
                    break
            if arquivo_encontrado:
                break

        if not arquivo_encontrado or not arquivo_encontrado.exists():
            resultado["message"] = f"Arquivo não encontrado: {nome_arquivo}"
            return resultado

        resultado["source"] = str(arquivo_encontrado)

        # Monta o caminho de destino baseado no modo
        nome_limpo = sanitizar(arquivo_encontrado.name)
        artista_limpo = sanitizar(artista)
        musica_limpo = sanitizar(musica)

        if organizacao_modo == "flat":
            # Flat: move diretamente para a raiz (sem subpasta)
            destino = DOWNLOADS_DIR / nome_limpo
        elif organizacao_modo == "pasta_unica":
            nome_pasta = f"{artista_limpo} - {musica_limpo}"
            destino = DOWNLOADS_DIR / nome_pasta / nome_limpo
        elif organizacao_modo == "artista":
            destino = DOWNLOADS_DIR / artista_limpo / musica_limpo / nome_limpo
        elif organizacao_modo == "pasta_personalizada":
            nome_pasta = sanitizar(pasta_personalizada) if pasta_personalizada else "Downloads"
            destino = DOWNLOADS_DIR / nome_pasta / nome_limpo
        else:
            # Fallback: raiz
            destino = DOWNLOADS_DIR / nome_limpo

        # Cria as pastas necessárias
        destino.parent.mkdir(parents=True, exist_ok=True)

        # Move o arquivo
        if arquivo_encontrado != destino:
            shutil.move(str(arquivo_encontrado), str(destino))

        resultado["destination"] = str(destino)
        resultado["success"] = True
        resultado["message"] = f"Movido para: {destino.relative_to(DOWNLOADS_DIR)}"

        # Limpa pastas vazias que sobrarem
        limpar_pastas_vazias(arquivo_encontrado.parent)

        return resultado

    except Exception as e:
        resultado["message"] = f"Erro ao mover: {e}"
        return resultado


def limpar_pastas_vazias(caminho: Path):
    """Remove pastas vazias recursivamente a partir do caminho fornecido."""
    try:
        while caminho and caminho != DOWNLOADS_DIR and caminho.exists():
            if caminho.is_dir() and not any(caminho.iterdir()):
                caminho.rmdir()
                caminho = caminho.parent
            else:
                break
    except Exception:
        pass


def montar_caminho_arquivo(artista, musica, nome_arq, organizacao_modo, pasta_personalizada="", pasta_base=""):
    """
    Monta o caminho completo do arquivo baseado no modo de organização.

    Modos disponíveis:
        - "flat": Tudo diretamente na pasta base (ou raiz downloads)
        - "pasta_unica": "Artista - Musica/arquivo.mp3"
        - "artista": "Artista/Musica/arquivo.mp3"
        - "pasta_personalizada": "nome_escolhido/arquivo.mp3"
    """
    nome_limpo = sanitizar(nome_arq)
    artista_limpo = sanitizar(artista)
    musica_limpo = sanitizar(musica)

    if pasta_base:
        base_path = pasta_base.strip('/')
    else:
        base_path = ""

    if organizacao_modo == "flat":
        if base_path:
            caminho_final = f"{base_path}/{nome_limpo}"
        else:
            caminho_final = nome_limpo

    elif organizacao_modo == "pasta_unica":
        nome_pasta = f"{artista_limpo} - {musica_limpo}"
        if base_path:
            caminho_final = f"{base_path}/{nome_pasta}/{nome_limpo}"
        else:
            caminho_final = f"{nome_pasta}/{nome_limpo}"

    elif organizacao_modo == "artista":
        if base_path:
            caminho_final = f"{base_path}/{artista_limpo}/{musica_limpo}/{nome_limpo}"
        else:
            caminho_final = f"{artista_limpo}/{musica_limpo}/{nome_limpo}"

    elif organizacao_modo == "pasta_personalizada":
        nome_pasta = sanitizar(pasta_personalizada) if pasta_personalizada else "Downloads"
        if base_path:
            caminho_final = f"{base_path}/{nome_pasta}/{nome_limpo}"
        else:
            caminho_final = f"{nome_pasta}/{nome_limpo}"
    else:
        if base_path:
            caminho_final = f"{base_path}/{nome_limpo}"
        else:
            caminho_final = nome_limpo

    return caminho_final.lstrip("/")


def baixar_arquivo(session, usuario, arquivo, tamanho, artista, musica, organizacao_modo, pasta_base="", pasta_personalizada="") -> bool:
    """
    Envia o download para o slskd com o caminho de destino correto.

    organizacao_modo:
        - "flat": Tudo na mesma pasta (sem subpastas)
        - "pasta_unica": Uma pasta por música "Artista - Musica/"
        - "artista": Pasta do artista, depois pasta da música "Artista/Musica/"
        - "pasta_personalizada": Nome de pasta personalizado
    """
    try:
        nome_arq = arquivo.replace("\\", "/").split("/")[-1]

        # O slskd não suporta localFilename para controlar destino.
        # A organização é feita via pós-processamento em monitorar_downloads_slskd.
        payload = {
            "filename": arquivo,
            "size": tamanho,
        }

        # O caminho informado aqui é apenas para logs - o slskd ignora
        push_log(f"📁 Organizacao: {organizacao_modo} (pós-processamento ativo)", "info")

        r = session.post(
            f"{SLSKD_URL}/api/v0/transfers/downloads/{usuario}",
            json=[payload],
            timeout=15,
        )

        if r.status_code in (200, 201, 204):
            push_log(f"✅ Download enfileirado: {nome_arq}", "info")
            return True
        else:
            push_log(f"❌ Erro slskd: {r.status_code} - {r.text[:200]}", "error")
            return False

    except Exception as e:
        push_log(f"❌ Falha download: {e}", "error")
        return False

def atualizar_musica(idx: int, **kwargs):
    for k, v in kwargs.items():
        state["musicas"][idx][k] = v

def monitorar_downloads_slskd():
    """Thread que monitora o slskd e atualiza status pra 'baixado' quando termina de verdade"""
    session = criar_sessao()
    if not obter_token(session):
        push_log("❌ Monitor: Falha ao autenticar no slskd", "error")
        return

    while state["running"] or any(m["status"] in ["enfileirado", "baixando"] for m in state["musicas"]):
        try:
            r = session.get(f"{SLSKD_URL}/api/v0/transfers/downloads", timeout=10)
            if r.status_code == 200:
                downloads_slskd = r.json()
                
                for grupo in downloads_slskd:
                    for dl in grupo.get("directories", []):
                        for arquivo in dl.get("files", []):
                            usuario_slskd = grupo.get("username", "")
                            tamanho_slskd = arquivo.get("size", 0)
                            nome_arq_slskd = arquivo["filename"].replace("\\", "/").split("/")[-1]
                            status_slskd = arquivo.get("state", "")
                            
                            for idx, m in enumerate(state["musicas"]):
                                if m["status"] not in ["enfileirado", "baixando"]:
                                    continue

                                # Variáveis para match
                                m_usuario = m.get("usuario", "")
                                m_tamanho = m.get("tamanho_bytes", 0)

                                is_match = False
                                if m_usuario and m_tamanho:
                                    is_match = (m_usuario == usuario_slskd and m_tamanho == tamanho_slskd)

                                if not is_match:
                                    nome_musica = sanitizar(m["musica"]).lower()
                                    nome_artista = sanitizar(m["artista"]).lower()
                                    nome_slskd = nome_arq_slskd.lower()
                                    is_match = (nome_musica in nome_slskd or nome_artista in nome_slskd)

                                if not is_match:
                                    # Log de debug: mostra por que não houve match
                                    push_log(f"   [DEBUG] slskd: {nome_arq_slskd[:50]} | musica: {m['musica'][:30]} - não bateu", "debug")

                                if is_match:
                                    if "Completed" in status_slskd and "Succeeded" in status_slskd:
                                        if m["status"] != "baixado":  # Evita contar 2x
                                            atualizar_musica(idx, status="baixado", progresso=100)
                                            state["baixados"] += 1
                                            push_log(f"✅ Download concluído: {m['musica']}")

                                            # Pós-processamento: mover arquivo para destino correto
                                            organizacao_modo = state["config"].get("organizacao_modo", "flat")
                                            pasta_personalizada = state["config"].get("pasta_personalizada", "")

                                            resultado_mover = mover_arquivo_para_destino(
                                                nome_arq_slskd,
                                                m.get("artista", ""),
                                                m.get("musica", ""),
                                                organizacao_modo,
                                                pasta_personalizada
                                            )

                                            if resultado_mover["success"]:
                                                push_log(f"📁 {resultado_mover['message']}", "info")
                                            else:
                                                push_log(f"⚠️ Pós-processamento: {resultado_mover['message']}", "warn")
                                    
                                    elif "Rejected" in status_slskd:
                                        retries = m.get("retries", 0)
                                        if retries < 3:
                                            if "rejected_candidates" not in state:
                                                state["rejected_candidates"] = set()
                                            state["rejected_candidates"].add((m.get("usuario"), m.get("arquivo_original")))
                                            
                                            atualizar_musica(idx, status="pendente", retries=retries + 1)
                                            push_log(f"⚠️ Rejeitado por {m['usuario']}, tentando novamente (Tentativa {retries + 1}/3): {m['musica']}", "warn")
                                            
                                            transfer_id = arquivo.get("id")
                                            if transfer_id:
                                                try:
                                                    session.delete(f"{SLSKD_URL}/api/v0/transfers/downloads/{usuario_slskd}/{transfer_id}", timeout=10)
                                                except Exception as ex:
                                                    push_log(f"⚠️ Erro ao remover transfer do slskd: {ex}", "warn")
                                        else:
                                            if m["status"] != "erro":
                                                atualizar_musica(idx, status="erro")
                                                state["erros"] += 1
                                                push_log(f"❌ Falhou no slskd (Excedeu retries): {m['musica']}", "error")
                                                
                                    elif "Failed" in status_slskd or "TimedOut" in status_slskd:
                                        if m["status"] != "erro":
                                            atualizar_musica(idx, status="erro")
                                            state["erros"] += 1
                                            push_log(f"❌ Falhou no slskd: {m['musica']}", "error")
                                    
                                    elif "InProgress" in status_slskd:
                                        pct = arquivo.get("percentComplete", 0)
                                        atualizar_musica(idx, status="baixando", progresso=round(pct, 1))
                                        
        except Exception as e:
            push_log(f"⚠️ Erro no monitor: {e}", "warn")
        
        time.sleep(5)

    push_log("🏁 Monitor de downloads encerrado")


def run_downloads():
    """Thread principal de downloads."""
    cfg        = state["config"]
    formatos   = cfg["formatos"]
    bitrate_min= cfg["bitrate_minimo"]
    tempo_busca= cfg["tempo_busca"]
    tamanho_min= cfg["tamanho_minimo"]
    organizacao_modo = cfg.get("organizacao_modo", "flat")
    pasta_base = cfg.get("pasta_download", "")
    pasta_personalizada = cfg.get("pasta_personalizada", "")
    prioridade = {f: len(formatos) - i for i, f in enumerate(formatos)}

    # LIMPAR rejected_candidates no início de cada sessão
    # (o slskd rejeita por motivos temporários como usuário offline, etc.)
    if "rejected_candidates" in state:
        qtd = len(state["rejected_candidates"])
        if qtd > 0:
            push_log(f"🧹 Limpando {qtd} candidato(s) rejeitado(s) de sessões anteriores", "info")
            state["rejected_candidates"] = set()

    session = criar_sessao()
    if not obter_token(session):
        state["running"] = False
        push_log("❌ Falha ao autenticar no slskd", "error")
        return

    push_log("✅ Conectado ao slskd")

    # Inicia o monitor em paralelo
    monitor_thread = threading.Thread(target=monitorar_downloads_slskd, daemon=True)
    monitor_thread.start()

    push_log(f"📁 Modo de organização: {organizacao_modo}", "info")
    if organizacao_modo == "pasta_personalizada" and pasta_personalizada:
        push_log(f"📁 Pasta personalizada: {pasta_personalizada}", "info")

    while state["running"]:
        if state["stop_flag"]:
            push_log("⏹ Download interrompido pelo usuário", "warn")
            break

        while state["paused"] and not state["stop_flag"]:
            time.sleep(0.5)

        idx = next((i for i, m in enumerate(state["musicas"]) if m["status"] == "pendente"), None)
        if idx is None:
            if any(m["status"] in ["enfileirado", "baixando", "buscando"] for m in state["musicas"]):
                time.sleep(2)
                continue
            else:
                break

        m      = state["musicas"][idx]
        musica = m["musica"]
        artista= m["artista"]
        label  = f"{artista} — {musica}"

        state["atual"] = label
        atualizar_musica(idx, status="buscando")
        push_log(f"\n[{idx+1}/{state['total']}] {label}")

        resultados = buscar_musica(session, musica, artista, tempo_busca)
        candidatos = filtrar_candidatos(resultados, formatos, bitrate_min, tamanho_min, prioridade)

        if not candidatos:
            atualizar_musica(idx, status="não encontrado")
            state["erros"]    += 1
            state["pendentes"] = max(0, state["pendentes"] - 1)
            push_log(f"⚠️ Nenhum candidato válido para: {label}", "warn")
            continue

        melhor   = candidatos[0]
        duracao  = melhor.get("duracao", 0)
        razao    = melhor.get("razao", 0)
        dur_str  = f"{int(duracao)//60}:{int(duracao)%60:02d}" if duracao else "?:??"
        razao_str= f"{razao:.2f} MB/min" if razao else "N/A"

        push_log(
            f"✅ {melhor['formato'].upper()} | "
            f"{melhor['bitrate']}kbps | "
            f"{melhor['tamanho']/1e6:.1f}MB | "
            f"{dur_str} | {razao_str}"
        )

        atualizar_musica(idx,
            status   = "baixando",
            formato  = melhor["formato"].upper(),
            bitrate  = melhor["bitrate"],
            tamanho  = round(melhor["tamanho"] / 1e6, 1),
            tamanho_bytes = melhor["tamanho"],
            usuario  = melhor["usuario"],
            arquivo_original = melhor["arquivo"],
            score    = round(melhor["score"], 1),
        )

        ok = baixar_arquivo(session, melhor["usuario"], melhor["arquivo"],
                            melhor["tamanho"], artista, musica,
                            organizacao_modo, pasta_base, pasta_personalizada)

        if ok:
            atualizar_musica(idx, status="enfileirado")
            nome_arq = melhor['arquivo'].replace("\\", "/").split("/")[-1]
            push_log(f"⬇️ Enfileirado: {nome_arq}")
            state["pendentes"] = max(0, state["pendentes"] - 1)
            # NÃO soma em baixados aqui!
        else:
            atualizar_musica(idx, status="erro")
            state["erros"]     += 1
            state["pendentes"]  = max(0, state["pendentes"] - 1)
            push_log(f"❌ Falha no download: {label}", "error")

        time.sleep(cfg.get("espera_download", 5))

    state["running"]    = False
    state["stop_flag"]  = False
    state["atual"]      = ""
    push_log("🏁 Pipeline concluído!")


# ─── Rotas API ────────────────────────────────────────────────

@app.route('/api/status')
def api_status():
    musicas = state["musicas"]
    
    baixados = [m for m in musicas if m.get('status') == 'baixado']
    erros = [m for m in musicas if m.get('status') in ['erro', 'nao_encontrado', 'não encontrado']]
    pendentes = [m for m in musicas if m.get('status') in ['pendente', 'buscando']]
    enfileirados = [m for m in musicas if m.get('status') == 'enfileirado']
    baixando = [m for m in musicas if m.get('status') == 'baixando']
    
    total = len(musicas)
    
    return jsonify({
        "baixados": len(baixados),
        "erros": len(erros), 
        "pendentes": len(pendentes),
        "enfileirados": len(enfileirados),
        "baixando": len(baixando),
        "total": total,
        "musicas": musicas,
        "atual": state.get("atual", ""),
        "running": state.get("running", False),
        "paused": state.get("paused", False),
        "sucesso": round(len(baixados) / total * 100, 0) if total > 0 else 0
    })

@app.route("/api/load", methods=["POST"])
def api_load():
    data    = request.json or {}
    print(f"[DEBUG] /api/load recebido: {data}")
    caminho = data.get("arquivo", "musicas.xlsx")
    col_m   = data.get("col_musica", "Song")
    col_a   = data.get("col_artista", "Artist")
    print(f"[DEBUG] arquivo={caminho} col_m={col_m} col_a={col_a}")

    try:
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
        cab = {str(c.value).strip().lower(): c.column - 1 for c in ws[1] if c.value}

        if col_m.lower() not in cab:
            return jsonify({"ok": False, "erro": f"Coluna '{col_m}' não encontrada. Disponíveis: {list(cab.keys())}"}), 400
        if col_a.lower() not in cab:
            return jsonify({"ok": False, "erro": f"Coluna '{col_a}' não encontrada. Disponíveis: {list(cab.keys())}"}), 400

        idx_m, idx_a = cab[col_m.lower()], cab[col_a.lower()]
        musicas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            m = str(row[idx_m]).strip() if row[idx_m] else ""
            a = str(row[idx_a]).strip() if row[idx_a] else ""
            if m and m.lower() != "none":
                musicas.append({"musica": m, "artista": a, "status": "pendente",
                                "formato": "", "bitrate": 0, "tamanho": 0,
                                "usuario": "", "score": 0})

        state["musicas"]   = musicas
        state["total"]     = len(musicas)
        state["baixados"]  = 0
        state["erros"]     = 0
        state["pendentes"] = len(musicas)
        state["running"]   = False

        return jsonify({"ok": True, "total": len(musicas)})

    except FileNotFoundError:
        return jsonify({"ok": False, "erro": f"Arquivo não encontrado: {caminho}"}), 404
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/start", methods=["POST"])
def api_start():
    if state["running"]:
        return jsonify({"ok": False, "erro": "Já em execução"}), 400
    if not state["musicas"]:
        return jsonify({"ok": False, "erro": "Carregue um Excel primeiro"}), 400

    state["running"]   = True
    state["paused"]    = False
    state["stop_flag"] = False

    t = threading.Thread(target=run_downloads, daemon=True)
    t.start()
    return jsonify({"ok": True})


@app.route("/api/pause", methods=["POST"])
def api_pause():
    state["paused"] = not state["paused"]
    return jsonify({"ok": True, "paused": state["paused"]})


@app.route("/api/stop", methods=["POST"])
def api_stop():
    state["stop_flag"] = True
    return jsonify({"ok": True})


@app.route("/api/clear-errors", methods=["POST"])
def api_clear_errors():
    erros_status = ["erro", "não encontrado", "nao encontrado", "nao_encontrado"]
    musicas_filtradas = [m for m in state["musicas"] if m.get("status") not in erros_status]
    quantidade_removida = len(state["musicas"]) - len(musicas_filtradas)
    
    state["musicas"] = musicas_filtradas
    state["total"] = len(musicas_filtradas)
    
    baixados = [m for m in musicas_filtradas if m.get("status") == "baixado"]
    pendentes = [m for m in musicas_filtradas if m.get("status") in ["pendente", "buscando", "enfileirado", "baixando"]]
    
    state["baixados"] = len(baixados)
    state["erros"] = 0
    state["pendentes"] = len(pendentes)
    
    push_log(f"🧹 Limpeza: {quantidade_removida} música(s) com erro removidas.", "info")
    return jsonify({"ok": True, "removidos": quantidade_removida})


@app.route("/api/config", methods=["POST"])
def api_config():
    data = request.json or {}
    cfg  = state["config"]
    if "formatos"              in data: cfg["formatos"]              = data["formatos"]
    if "bitrate_minimo"        in data: cfg["bitrate_minimo"]        = int(data["bitrate_minimo"])
    if "tempo_busca"           in data: cfg["tempo_busca"]           = max(60, int(data["tempo_busca"]))
    if "tamanho_minimo"        in data: cfg["tamanho_minimo"]        = int(data["tamanho_minimo"])
    if "organizacao_modo"      in data: cfg["organizacao_modo"]      = data["organizacao_modo"]
    if "pasta_download"        in data: cfg["pasta_download"]        = str(data["pasta_download"]).strip()
    if "pasta_personalizada"   in data: cfg["pasta_personalizada"]   = str(data["pasta_personalizada"]).strip()
    return jsonify({"ok": True, "config": cfg})


@app.route("/api/export/not-found", methods=["POST"])
def api_export_not_found():
    """Exporta apenas as músicas não encontradas para Excel"""
    try:
        musicas_nao_encontradas = [m for m in state["musicas"] if m["status"] == "não encontrado"]

        if not musicas_nao_encontradas:
            return jsonify({"ok": False, "erro": "Nenhuma música não encontrada para exportar"}), 400

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Não Encontrados"

        headers = ["Música", "Artista", "Status"]
        ws.append(headers)

        for m in musicas_nao_encontradas:
            ws.append([m["musica"], m["artista"], m["status"]])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=nao_encontrados.xlsx"}
        )

    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/fix-permissions", methods=["POST"])
def api_fix_permissions():
    """Executa o script remove_restrictions_download.py para corrigir permissões"""
    try:
        import subprocess

        script_path = "/workspace/user_input_files/remove_restrictions_download.py"

        if not os.path.exists(script_path):
            return jsonify({"ok": False, "erro": f"Script não encontrado: {script_path}"}), 404

        result = subprocess.run(
            ["sudo", "python3", script_path],
            capture_output=True,
            text=True,
            timeout=120
        )

        if result.returncode == 0:
            push_log("✅ Permissões corrigidas com sucesso!", "info")
            return jsonify({
                "ok": True,
                "mensagem": "Permissões corrigidas!",
                "output": result.stdout
            })
        else:
            push_log(f"⚠️ Erro ao corrigir permissões: {result.stderr}", "warn")
            return jsonify({
                "ok": False,
                "erro": result.stderr or "Erro desconhecido",
                "output": result.stdout
            }), 500

    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "erro": "Tempo limite excedido"}), 500
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/logs")
def api_logs_sse():
    def generate():
        while True:
            try:
                item = log_queue.get(timeout=1)
                yield f"data: {json.dumps(item)}\n\n"
            except Empty:
                yield ": ping\n\n"
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/")
def index():
    response = send_from_directory(".", "index.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", default=8082, type=int)
    args = parser.parse_args()

    print(f"\n🎵 Interface web iniciada → http://localhost:{args.port}\n")
    app.run(host="0.0.0.0", port=args.port, debug=False, threaded=True)
