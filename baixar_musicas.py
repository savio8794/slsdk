#!/usr/bin/env python3
"""
Baixador Automático de Músicas via slskd
========================================

Melhorias implementadas:
- Busca inteligente progressiva
- Cache de músicas já processadas
- Espera inteligente por resultados
- Prioridade dinâmica de formatos
- Filtro de arquivos suspeitos
- Organização automática por artista/música
- Retry automático em falhas HTTP
- Session reutilizável com pool
- Logs mais limpos
- Controle de duplicatas
- Melhor score de qualidade
- Timeout configurável
- Sanitização de nomes
"""

import argparse
import json
import logging
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ============================================================
# CONFIGURAÇÕES
# ============================================================

SLSKD_URL = "http://localhost:5030"
SLSKD_USUARIO = "slskd"
SLSKD_SENHA = "slskd"

EXCEL_ARQUIVO = "musicas.xlsx"

COLUNA_MUSICA = "musica"
COLUNA_ARTISTA = "artista"

FORMATOS_PADRAO = ["flac", "wav", "mp3"]

BITRATE_MINIMO_MP3 = 192
TAMANHO_MINIMO = 7_000_000

TEMPO_MAXIMO_BUSCA = 10
INTERVALO_BUSCA = 1
MINIMO_RESULTADOS_BUSCA = 5
ESPERA_DOWNLOAD_SEG = 5

LOG_DIR = Path("logs")


# ============================================================


def configurar_logs():

    LOG_DIR.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    log_file = LOG_DIR / f"downloads_{timestamp}.log"

    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )

    logger = logging.getLogger("slskd-auto")

    return logger, log_file


def criar_sessao():

    session = requests.Session()

    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )

    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=10,
        pool_maxsize=10,
    )

    session.mount("http://", adapter)
    session.mount("https://", adapter)

    return session


def obter_token(session, log):

    try:

        r = session.post(
            f"{SLSKD_URL}/api/v0/session",
            json={
                "username": SLSKD_USUARIO,
                "password": SLSKD_SENHA,
            },
            timeout=10,
        )

        if r.status_code == 200:

            token = r.json().get("token")

            if token:

                session.headers.update({
                    "Authorization": f"Bearer {token}"
                })

                log.info("✅ Autenticado no slskd")

                return True

        log.error(f"Falha autenticação: {r.status_code}")

    except Exception as e:
        log.error(f"Erro conexão slskd: {e}")

    return False


def limpar_termo(texto: str) -> str:

    texto = re.sub(r"\(.*?\)", "", texto)
    texto = re.sub(r"\[.*?\]", "", texto)

    texto = re.sub(
        r"\b(feat|ft|featuring)\.?\b.*",
        "",
        texto,
        flags=re.IGNORECASE,
    )

    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def sanitizar_pasta(texto: str) -> str:

    texto = re.sub(r'[<>:"/\\|?*]', "", texto)

    return texto.strip()


def fazer_busca(session, termo, log):

    try:

        r = session.post(
            f"{SLSKD_URL}/api/v0/searches",
            json={"searchText": termo},
            timeout=10,
        )

        if r.status_code == 409:

            log.debug("Busca duplicada detectada")

            r_list = session.get(
                f"{SLSKD_URL}/api/v0/searches",
                timeout=10,
            )

            if r_list.status_code == 200:

                for busca in r_list.json():

                    if (
                        busca.get("searchText", "").lower()
                        == termo.lower()
                    ):

                        search_id = busca.get("id")

                        return aguardar_resultados(
                            session,
                            search_id,
                            log
                        )

            return []

        if r.status_code not in (200, 201):

            log.error(f"Erro busca: {r.status_code}")

            return []

        search_id = r.json().get("id")

        if not search_id:

            log.error("Search ID inválido")

            return []

        resultados = aguardar_resultados(session, search_id, log)

        # Deleta a busca para não entupir a fila nas próximas execuções
        try:
            session.delete(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
            log.debug(f"Busca {search_id} deletada da fila")
        except Exception:
            pass

        return resultados

    except Exception as e:

        log.error(f"Erro busca: {e}")

        return []


TIMEOUT_BUSCA_ABSOLUTO = 60   # segundos máximos esperando peers
TIMEOUT_ESTABILIDADE  = 8    # segundos sem crescimento → considera estável
INTERVALO_POLLING     = 2    # segundos entre cada consulta


def aguardar_resultados(session, search_id, log):
    """
    Aguarda a busca do slskd consultando GET /api/v0/searches/{id}
    e inspecionando 'isComplete' + 'responses' dentro do objeto retornado.

    O endpoint /searches/{id}/responses NÃO existe em todas as versões;
    a API correta embute as respostas dentro do objeto principal da busca.
    """

    inicio       = time.time()
    ultimo_total = 0
    ultimo_cresc = time.time()

    # ── Fase 1: aguarda sair do estado Queued ──────────────
    log.info("⏳ Aguardando busca sair da fila...")
    while True:
        if time.time() - inicio > 60:
            log.warning("⌛ Timeout aguardando fila — busca nunca saiu de Queued")
            return []
        try:
            r = session.get(f"{SLSKD_URL}/api/v0/searches/{search_id}", timeout=10)
            if r.status_code == 200:
                state = r.json().get("state", "")
                if state != "Queued":
                    log.info(f"▶ Busca iniciada (state={state})")
                    break
                log.debug(f"Ainda na fila... ({time.time()-inicio:.0f}s)")
        except Exception as e:
            log.debug(f"Erro aguardando fila: {e}")
        time.sleep(2)

    # ── Fase 2: polling de resultados ──────────────────────
    while True:

        agora   = time.time()
        elapsed = agora - inicio

        try:
            r = session.get(
                f"{SLSKD_URL}/api/v0/searches/{search_id}",
                timeout=10,
            )

            if r.status_code == 200:
                dados     = r.json()
                respostas = dados.get("responses") or []
                is_done   = dados.get("isComplete", False)
                state     = dados.get("state", "")

                total = sum(len(p.get("files", [])) for p in respostas)

                if total != ultimo_total:
                    log.info(
                        f"📦 Resultados: {total} arquivos "
                        f"(+{total - ultimo_total}) [{elapsed:.0f}s]"
                    )
                    ultimo_total = total
                    ultimo_cresc = agora

                # Busca concluída pelo próprio slskd
                if is_done or state in ("Completed", "TimedOut"):
                    log.info(
                        f"✅ Busca concluída ({total} arquivos, "
                        f"{elapsed:.0f}s, state={state})"
                    )
                    return respostas

                # Estabilizou localmente com resultados suficientes
                if total >= 10 and (agora - ultimo_cresc) > TIMEOUT_ESTABILIDADE:
                    log.info(
                        f"✅ Estabilizou com {total} arquivos ({elapsed:.0f}s)"
                    )
                    return respostas

                # Timeout absoluto de segurança
                if elapsed > TIMEOUT_BUSCA_ABSOLUTO:
                    log.warning(
                        f"⌛ Timeout {TIMEOUT_BUSCA_ABSOLUTO}s — "
                        f"retornando {total} arquivo(s)"
                    )
                    return respostas

            elif r.status_code == 404:
                log.error(f"Search ID {search_id} não encontrado (404)")
                return []

            else:
                log.debug(f"Status inesperado: {r.status_code}")

        except Exception as e:
            log.debug(f"Erro ao consultar busca: {e}")

        time.sleep(INTERVALO_POLLING)



def buscar_musica(session, musica, artista, log):

    musica_limpa = limpar_termo(musica)
    artista_limpo = limpar_termo(artista)

    # Primeiro artista apenas
    primeiro_artista = (
        artista_limpo
        .split(",")[0]
        .split("&")[0]
        .split(";")[0]
        .strip()
    )

    tentativas = [

        # 1 — Busca normal completa
        f"{artista_limpo} {musica_limpa}",

        # 2 — Primeiro artista + música
        f"{primeiro_artista} {musica_limpa}",

        # 4 — Artista + primeiras palavras
        f"{primeiro_artista} {' '.join(musica_limpa.split()[:3])}",
    ]

    # Remove duplicatas mantendo ordem
    vistas = set()

    tentativas = [
        t for t in tentativas
        if not (
            t.lower() in vistas
            or vistas.add(t.lower())
        )
    ]

    melhores_resultados = []

    for i, termo in enumerate(tentativas, 1):

        log.info(f"🔍 Tentativa {i}: {termo}")

        resultados = fazer_busca(
            session,
            termo,
            log
        )

        total = sum(
            len(r.get("files", []))
            for r in resultados
        )

        if total > 0:

            log.info(f"📦 {total} arquivo(s) encontrado(s) na tentativa {i}")

            # guarda o melhor resultado até agora
            if total > sum(
                len(r.get("files", []))
                for r in melhores_resultados
            ):
                melhores_resultados = resultados

        else:

            log.warning(f"Nenhum resultado para tentativa {i}: '{termo}'")

    total_final = sum(len(r.get("files", [])) for r in melhores_resultados)
    if total_final == 0:
        log.warning("⚠️ NENHUM resultado em nenhuma tentativa!")
    else:
        log.info(f"📦 Melhor tentativa retornou {total_final} arquivo(s)")

    # retorna o melhor conjunto encontrado
    return melhores_resultados
    

def calcular_score(candidato):

    score = 0

    formato = candidato["formato"]
    bitrate = candidato["bitrate"]
    tamanho = candidato["tamanho"]

    # prioridade do formato
    score += candidato["prioridade"] * 10000

    # FLAC sempre priorizado
    if formato == "flac":
        score += 5000

    # WAV
    elif formato == "wav":
        score += 3000

    # MP3
    elif formato == "mp3":

        # bitrate válido
        if bitrate and bitrate > 0:

            score += bitrate * 5

        else:

            # VBR/desconhecido
            score += 1000

    # tamanho ajuda muito
    score += tamanho / 1_000_000

    return score


def filtrar_candidatos(
    resultados,
    formatos,
    prioridade,
    log
):

    candidatos = []

    for resposta in resultados:

        usuario = resposta.get("username", "")

        for arq in resposta.get("files", []):


            nome = arq.get("filename", "")

            ext = (
                nome.rsplit(".", 1)[-1].lower()
                if "." in nome else ""
            )

            bitrate_raw = (
                arq.get("bitRate")
                or arq.get("bitrate")
                or 0
            )

            try:
                bitrate = int(float(bitrate_raw))
            except:
                bitrate = 0

            tamanho = arq.get("size", 0) or 0
            duracao = (
                arq.get("length")
                or arq.get("duration")
                or 0
            )

            nome_lower = nome.lower()

            log.debug(
                f"ANALISANDO | "
                f"ext={ext} | "
                f"bitrate={bitrate} | "
                f"size={tamanho/1e6:.1f}MB | "
                f"nome={nome}"
            )

            if ext not in formatos:

                log.debug(
                    f"DESCARTADO FORMATO | "
                    f"ext={ext} | "
                    f"{nome}"
                )

                continue

            if tamanho < TAMANHO_MINIMO:

                log.debug(
                    f"DESCARTADO TAMANHO | "
                    f"size={tamanho/1e6:.1f}MB | "
                    f"{nome}"
                )

                continue

            if ext == "mp3":

                # bitrate desconhecido → aceita
                if bitrate <= 0:

                    bitrate = 320

                # rejeita somente muito ruins
                elif bitrate < 192:

                    log.debug(
                        f"DESCARTADO BITRATE | "
                        f"{bitrate}kbps | "
                        f"{nome}"
                    )

                    continue

            candidato = {
                "usuario": usuario,
                "arquivo": nome,
                "formato": ext,
                "bitrate": bitrate,
                "tamanho": tamanho,
                "duracao": duracao,
                "prioridade": prioridade.get(ext, 0),
            }

            candidato["score"] = calcular_score(candidato)

            candidatos.append(candidato)

    candidatos.sort(
        key=lambda x: x["score"],
        reverse=True,
    )

    log.info(f"🎯 {len(candidatos)} candidato(s) válido(s)")

    return candidatos


def baixar_arquivo(
    session,
    usuario,
    arquivo,
    tamanho,
    artista,
    musica,
    organizacao_modo,
    pasta_personalizada,
    log
):
    """
    Envia o download para o slskd com o caminho de destino correto.

    organizacao_modo:
        - "flat": Tudo na mesma pasta (sem subpastas)
        - "pasta_unica": Uma pasta por música "Artista - Musica/"
        - "artista": Pasta do artista, depois pasta da música "Artista/Musica/"
        - "pasta_personalizada": Nome de pasta personalizado
    """
    try:
        nome_arq = arquivo.replace("\\", "/").split("/")[-1]
        nome_limpo = sanitizar_pasta(nome_arq)
        artista_limpo = sanitizar_pasta(artista)
        musica_limpo = sanitizar_pasta(musica)

        # Monta o caminho baseado no modo de organização
        if organizacao_modo == "flat":
            # Modo flat: todos os arquivos diretamente na pasta base (sem subpastas)
            caminho_final = nome_limpo

        elif organizacao_modo == "pasta_unica":
            # Uma pasta por música: "Artista - Musica/arquivo.mp3"
            nome_pasta = f"{artista_limpo} - {musica_limpo}"
            caminho_final = f"{nome_pasta}/{nome_limpo}"

        elif organizacao_modo == "artista":
            # Artista/Música: "Artista/Musica/arquivo.mp3"
            caminho_final = f"{artista_limpo}/{musica_limpo}/{nome_limpo}"

        elif organizacao_modo == "pasta_personalizada":
            # Nome de pasta personalizado
            nome_pasta = sanitizar_pasta(pasta_personalizada) if pasta_personalizada else "Downloads"
            caminho_final = f"{nome_pasta}/{nome_limpo}"
        else:
            # Fallback para flat
            caminho_final = nome_limpo

        log.info(f"💾 Salvando em: {caminho_final}")

        # Estrutura correta para a API do slskd
        payload = {
            "filename": arquivo,
            "size": tamanho,
            "localFilename": caminho_final
        }

        r = session.post(
            f"{SLSKD_URL}/api/v0/transfers/downloads/{usuario}",
            json=[payload],
            timeout=15,
        )

        if r.status_code in (200, 201, 204):
            log.info(f"✅ Download enfileirado: {nome_limpo}")
            return True

        log.error(
            f"Erro download: {r.status_code} {r.text}"
        )

    except Exception as e:

        log.error(f"Falha download: {e}")

    return False


def ler_excel(
    caminho,
    col_musica,
    col_artista,
    log
):

    try:

        wb = openpyxl.load_workbook(caminho)

        ws = wb.active

        cabecalho = {
            str(cell.value).strip().lower():
            cell.column - 1
            for cell in ws[1]
            if cell.value
        }

        col_m = col_musica.lower().strip()
        col_a = col_artista.lower().strip()

        if col_m not in cabecalho:

            log.error(f"Coluna não encontrada: {col_musica}")

            sys.exit(1)

        if col_a not in cabecalho:

            log.error(f"Coluna não encontrada: {col_artista}")

            sys.exit(1)

        idx_m = cabecalho[col_m]
        idx_a = cabecalho[col_a]

        musicas = []

        for linha in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            musica = (
                str(linha[idx_m]).strip()
                if linha[idx_m]
                else ""
            )

            artista = (
                str(linha[idx_a]).strip()
                if linha[idx_a]
                else ""
            )

            if musica and musica.lower() != "none":

                musicas.append(
                    (musica, artista)
                )

        log.info(f"🎵 {len(musicas)} música(s) carregadas")

        return musicas

    except FileNotFoundError:

        log.error(f"Excel não encontrado: {caminho}")

        sys.exit(1)


def salvar_relatorio(
    log_file,
    baixados,
    nao_achados,
    erros
):

    relatorio = {
        "data": datetime.now().isoformat(),
        "baixados": baixados,
        "nao_achados": nao_achados,
        "erros": erros,
    }

    rel_path = log_file.with_suffix(".json")

    rel_path.write_text(
        json.dumps(
            relatorio,
            ensure_ascii=False,
            indent=2
        ),
        encoding="utf-8",
    )

    return rel_path


def main():

    parser = argparse.ArgumentParser(
        description="Downloader automático slskd"
    )

    parser.add_argument(
        "--excel",
        default=EXCEL_ARQUIVO,
    )

    parser.add_argument(
        "--coluna-musica",
        default=COLUNA_MUSICA,
    )

    parser.add_argument(
        "--coluna-artista",
        default=COLUNA_ARTISTA,
    )

    parser.add_argument(
        "--formatos",
        nargs="+",
        default=FORMATOS_PADRAO,
        choices=["flac", "wav", "mp3"],
    )

    parser.add_argument(
        "--organizacao",
        default="flat",
        choices=["flat", "pasta_unica", "artista", "pasta_personalizada"],
        help=(
            "Modo de organizacao das pastas:\n"
            "  flat               - Tudo na mesma pasta (sem subpastas)\n"
            "  pasta_unica        - Uma pasta por musica 'Artista - Musica/'\n"
            "  artista            - Pasta do artista 'Artista/Musica/'\n"
            "  pasta_personalizada - Nome de pasta personalizado\n"
        ),
    )

    parser.add_argument(
        "--pasta-personalizada",
        default="",
        help="Nome da pasta para o modo 'pasta_personalizada'",
    )

    args = parser.parse_args()

    prioridade = {
        formato: len(args.formatos) - i
        for i, formato in enumerate(args.formatos)
    }

    log, log_file = configurar_logs()

    log.info("=" * 60)
    log.info("🎵 SLSKD AUTO DOWNLOADER")
    log.info("=" * 60)
    log.info(f"📁 Modo de organizacao: {args.organizacao}")
    if args.organizacao == "pasta_personalizada" and args.pasta_personalizada:
        log.info(f"📁 Pasta personalizada: {args.pasta_personalizada}")

    musicas = ler_excel(
        args.excel,
        args.coluna_musica,
        args.coluna_artista,
        log
    )

    session = criar_sessao()

    if not obter_token(session, log):
        sys.exit(1)

    baixados = []
    nao_achados = []
    erros = []

    cache = set()

    total = len(musicas)

    for i, (musica, artista) in enumerate(musicas, 1):

        label = f"{artista} - {musica}"

        if label.lower() in cache:
            continue

        cache.add(label.lower())

        log.info(f"\n[{i}/{total}] {label}")

        resultados = buscar_musica(
            session,
            musica,
            artista,
            log
        )

        candidatos = filtrar_candidatos(
            resultados,
            args.formatos,
            prioridade,
            log
        )

        if not candidatos:

            nao_achados.append(label)

            log.warning("⚠️ Nenhum candidato válido")

            continue

        melhor = candidatos[0]

        nome_arquivo = (
            melhor["arquivo"]
            .replace("\\", "/")
            .split("/")[-1]
        )

        log.info(
            f"✅ "
            f"{melhor['formato'].upper()} | "
            f"{melhor['bitrate']}kbps | "
            f"{melhor['tamanho']/1e6:.1f}MB | "
            f"score={melhor['score']:.1f}"
        )

        ok = baixar_arquivo(
            session,
            melhor["usuario"],
            melhor["arquivo"],
            melhor["tamanho"],
            artista,
            musica,
            args.organizacao,
            args.pasta_personalizada,
            log
        )

        if ok:

            baixados.append({
                "musica": label,
                "arquivo": nome_arquivo,
                "usuario": melhor["usuario"],
                "formato": melhor["formato"],
                "score": melhor["score"],
            })

            log.info("⬇️ Download iniciado")

        else:

            erros.append(label)

        time.sleep(ESPERA_DOWNLOAD_SEG)

    log.info("\n" + "=" * 60)
    log.info("📊 RESUMO FINAL")
    log.info("=" * 60)

    log.info(f"✅ Baixados: {len(baixados)}")
    log.info(f"⚠️ Não encontrados: {len(nao_achados)}")
    log.info(f"❌ Erros: {len(erros)}")

    relatorio = salvar_relatorio(
        log_file,
        baixados,
        nao_achados,
        erros
    )

    log.info(f"📄 Relatório: {relatorio}")
    log.info("🌐 slskd: http://localhost:5030")


if __name__ == "__main__":
    main()
