#!/usr/bin/env python3
"""
Rekordbox — Verificador de Músicas Faltantes
==============================================
Compara seu Excel (lista desejada) com o XML exportado
do Rekordbox (biblioteca atual) e gera:

  - Relatório no terminal
  - logs/faltantes_YYYYMMDD.xlsx  ← pronto para usar no baixar_musicas.py
  - logs/faltantes_YYYYMMDD.json

Exportar o XML no Rekordbox:
  File → Export Collection in xml format → salve como rekordbox.xml

Uso:
    python rekordbox_faltantes.py
    python rekordbox_faltantes.py --xml minha_biblioteca.xml --excel musicas.xlsx
    python rekordbox_faltantes.py --similaridade 0.75   # match menos rigoroso
"""

import argparse
import json
import logging
import sys
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Tenta importar difflib para similaridade ──────────────
from difflib import SequenceMatcher

LOG_DIR = Path("logs")

# ============================================================
#  CONFIGURACOES
# ============================================================
XML_REKORDBOX   = "rekordbox.xml"
EXCEL_ARQUIVO   = "musicas.xlsx"
COLUNA_MUSICA   = "musica"
COLUNA_ARTISTA  = "artista"
SIMILARIDADE    = 0.82   # 0.0 a 1.0 — quão parecido precisa ser para considerar "encontrado"
# ============================================================


def configurar_logs():
    LOG_DIR.mkdir(exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = LOG_DIR / f"faltantes_{ts}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger("rekordbox"), log_file


def normalizar(texto: str) -> str:
    """Lowercase, sem acentos, sem espaços duplos."""
    texto = texto.lower().strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.split())


def similaridade(a: str, b: str) -> float:
    return SequenceMatcher(None, normalizar(a), normalizar(b)).ratio()


def ler_excel(caminho: str, col_musica: str, col_artista: str, log) -> list:
    """Lê o Excel e retorna lista de (musica, artista)."""
    try:
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active

        cabecalho = {
            str(cell.value).strip().lower(): cell.column - 1
            for cell in ws[1] if cell.value
        }

        col_m = col_musica.strip().lower()
        col_a = col_artista.strip().lower()

        if col_m not in cabecalho:
            log.error(f"Coluna '{col_musica}' não encontrada. Disponíveis: {list(cabecalho.keys())}")
            sys.exit(1)
        if col_a not in cabecalho:
            log.error(f"Coluna '{col_artista}' não encontrada. Disponíveis: {list(cabecalho.keys())}")
            sys.exit(1)

        idx_m = cabecalho[col_m]
        idx_a = cabecalho[col_a]

        musicas = []
        for linha in ws.iter_rows(min_row=2, values_only=True):
            m = str(linha[idx_m]).strip() if linha[idx_m] else ""
            a = str(linha[idx_a]).strip() if linha[idx_a] else ""
            if m and m.lower() != "none":
                musicas.append((m, a))

        log.info(f"Excel: {len(musicas)} música(s) lidas de '{caminho}'")
        return musicas

    except FileNotFoundError:
        log.error(f"Excel não encontrado: {caminho}")
        sys.exit(1)


def ler_rekordbox_xml(caminho: str, log) -> list:
    """
    Lê o XML exportado pelo Rekordbox e retorna lista de
    dicts com {nome, artista, album, caminho}.
    """
    try:
        tree = ET.parse(caminho)
        root = tree.getroot()
    except FileNotFoundError:
        log.error(f"XML não encontrado: {caminho}")
        log.error("Exporte pelo Rekordbox: File → Export Collection in xml format")
        sys.exit(1)
    except ET.ParseError as e:
        log.error(f"Erro ao ler XML: {e}")
        sys.exit(1)

    faixas = []
    # O XML do Rekordbox tem estrutura: DJ_PLAYLISTS > COLLECTION > TRACK
    colecao = root.find(".//COLLECTION")
    if colecao is None:
        log.error("Estrutura XML inesperada — COLLECTION não encontrada.")
        log.error("Certifique-se de exportar pelo menu File > Export Collection in xml format")
        sys.exit(1)

    for track in colecao.findall("TRACK"):
        faixas.append({
            "nome":    track.get("Name", ""),
            "artista": track.get("Artist", ""),
            "album":   track.get("Album", ""),
            "caminho": track.get("Location", ""),
        })

    log.info(f"Rekordbox XML: {len(faixas)} faixa(s) na biblioteca ('{caminho}')")
    return faixas


def encontrar_no_rekordbox(musica: str, artista: str, biblioteca: list, limiar: float) -> dict | None:
    """
    Tenta encontrar a música na biblioteca do Rekordbox.
    Retorna o match mais próximo se acima do limiar, ou None.
    """
    melhor_score = 0.0
    melhor_match = None

    for faixa in biblioteca:
        score_musica  = similaridade(musica,  faixa["nome"])
        score_artista = similaridade(artista, faixa["artista"]) if artista else 1.0
        # Média ponderada: nome tem peso maior
        score = (score_musica * 0.65) + (score_artista * 0.35)

        if score > melhor_score:
            melhor_score = score
            melhor_match = faixa

    if melhor_score >= limiar:
        return {"faixa": melhor_match, "score": round(melhor_score, 3)}
    return None


def salvar_excel_faltantes(faltantes: list, log_file: Path) -> Path:
    """Salva os faltantes num Excel pronto para usar no baixar_musicas.py."""
    caminho = log_file.with_suffix(".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Faltantes"

    # Cabeçalho
    ws.append(["musica", "artista"])
    for item in faltantes:
        ws.append([item["musica"], item["artista"]])

    # Largura das colunas
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 30

    wb.save(caminho)
    return caminho


def main():
    parser = argparse.ArgumentParser(
        description="Compara Excel com biblioteca Rekordbox e lista o que falta baixar."
    )
    parser.add_argument("--xml",            default=XML_REKORDBOX,  help="Caminho do XML exportado pelo Rekordbox")
    parser.add_argument("--excel",          default=EXCEL_ARQUIVO,  help="Caminho do arquivo Excel com wishlist")
    parser.add_argument("--coluna-musica",  default=COLUNA_MUSICA,  help="Nome da coluna de músicas no Excel")
    parser.add_argument("--coluna-artista", default=COLUNA_ARTISTA, help="Nome da coluna de artistas no Excel")
    parser.add_argument("--similaridade",   default=SIMILARIDADE, type=float,
                        help=f"Limiar de similaridade 0.0-1.0 (padrão: {SIMILARIDADE})")
    args = parser.parse_args()

    log, log_file = configurar_logs()

    log.info("=" * 55)
    log.info("🎛️  Rekordbox — Verificador de Faltantes")
    log.info(f"   Limiar de similaridade: {args.similaridade}")
    log.info("=" * 55)

    wishlist  = ler_excel(args.excel, args.coluna_musica, args.coluna_artista, log)
    biblioteca = ler_rekordbox_xml(args.xml, log)

    log.info(f"\nComparando {len(wishlist)} música(s) do Excel com {len(biblioteca)} faixa(s) do Rekordbox...\n")

    encontrados = []
    faltantes   = []
    parciais    = []   # match abaixo do limiar mas próximo (>0.6)

    for musica, artista in wishlist:
        label  = f"{artista} — {musica}"
        match  = encontrar_no_rekordbox(musica, artista, biblioteca, args.similaridade)

        if match:
            log.info(f"  ✅ {label}")
            log.info(f"     → '{match['faixa']['artista']} — {match['faixa']['nome']}' (score {match['score']})")
            encontrados.append({"musica": musica, "artista": artista, "match": match["faixa"], "score": match["score"]})
        else:
            # Verifica se tem algo próximo (para avisar)
            match_parcial = encontrar_no_rekordbox(musica, artista, biblioteca, 0.60)
            if match_parcial:
                log.warning(f"  ⚠️  {label}")
                log.warning(f"     → Possível match: '{match_parcial['faixa']['artista']} — {match_parcial['faixa']['nome']}' (score {match_parcial['score']}) — abaixo do limiar")
                parciais.append({"musica": musica, "artista": artista, "possivel": match_parcial["faixa"], "score": match_parcial["score"]})
                faltantes.append({"musica": musica, "artista": artista})
            else:
                log.warning(f"  ❌ FALTANDO: {label}")
                faltantes.append({"musica": musica, "artista": artista})

    # ── Resumo ──────────────────────────────────────────────
    log.info("\n" + "=" * 55)
    log.info("📊 RESUMO")
    log.info("=" * 55)
    log.info(f"  ✅ Já no Rekordbox:      {len(encontrados):>4}")
    log.info(f"  ⚠️  Match parcial (baixar): {len(parciais):>4}")
    log.info(f"  ❌ Faltando (baixar):    {len(faltantes):>4}")
    log.info(f"  📋 Total no Excel:       {len(wishlist):>4}")

    if faltantes:
        # Salva Excel de faltantes
        xlsx_path = salvar_excel_faltantes(faltantes, log_file)
        log.info(f"\n  📥 Excel de faltantes salvo em: {xlsx_path}")
        log.info("     Use para baixar automaticamente:")
        log.info(f"     python baixar_musicas.py --excel {xlsx_path.name}")

        # Salva JSON completo
        json_path = log_file.with_name(log_file.stem + "_resultado.json")
        json_path.write_text(
            json.dumps({
                "data": datetime.now().isoformat(),
                "encontrados": encontrados,
                "faltantes":   faltantes,
                "parciais":    parciais,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        log.info(f"  📄 Relatório JSON:          {json_path}")
    else:
        log.info("\n  🎉 Todas as músicas do Excel já estão no Rekordbox!")


if __name__ == "__main__":
    main()
